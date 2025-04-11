from io import BytesIO

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.cmdb.constants import ENUM, ORGANIZATION, USER


class Export:
    def __init__(self, attrs):
        self.attrs = attrs

    def set_row_color(self, sheet, row_num, color):
        """行添加颜色"""
        for cell in sheet[row_num]:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def generate_header(self):
        """创建Excel文件, 设置属性与样式"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.sheet_format.defaultColWidth = 20
        sheet.sheet_format.defaultRowHeight = 15
        attrs_name, attrs_id, index = [], [], 0

        for attr_info in self.attrs:
            attr_name = f'{attr_info["attr_name"]}(必填)' if attr_info.get("is_required") else attr_info["attr_name"]
            attrs_name.append(attr_name)
            attrs_id.append(attr_info["attr_id"])
            index += 1
            if attr_info["attr_type"] == ENUM:
                sheet.add_data_validation(
                    self.set_enum_validation_by_sheet_data(workbook, attr_info["attr_name"], attr_info["option"], index)
                )

        sheet.append(attrs_name)
        sheet.append(attrs_id)
        self.set_row_color(sheet, 1, "92D050")
        self.set_row_color(sheet, 2, "C6EFCE")

        return workbook

    def return_bytesio(self, workbook):
        """返回一个文件流"""
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return file_stream

    def set_enum_validation_by_sheet_data(self, workbook, filed_name, option, index):
        """设置枚举值, 通过sheet数据, 单选"""
        value_list = [i["name"] for i in option]

        # 将枚举数据放入sheet页
        filed_sheet = workbook.create_sheet(title=filed_name)
        for r, v in enumerate(value_list, start=1):
            filed_sheet.cell(row=r, column=1, value=v)

        # 创建 DataValidation 对象
        col = get_column_letter(index)
        last_row = len(filed_sheet["A"])
        dv = DataValidation(type="list", formula1=f"='{filed_sheet.title}'!$A$1:$A{last_row}")
        dv.sqref = f"{col}3:{col}999"

        return dv

    def export_template(self):
        """导出模板"""
        workbook = self.generate_header()
        return self.return_bytesio(workbook)

    def export_inst_list(self, inst_list):
        """导出实例列表"""
        workbook = self.generate_header()
        # 找出枚举属性
        enum_field_dict = {
            attr_info["attr_id"]: {i["id"]: i["name"] for i in attr_info["option"]}
            for attr_info in self.attrs
            if attr_info["attr_type"] in {ORGANIZATION, USER, ENUM}
        }
        for inst_info in inst_list:
            sheet_data = []
            for attr in self.attrs:
                if attr["attr_type"] in {ORGANIZATION, USER}:
                    sheet_data.append(
                        str([enum_field_dict[attr["attr_id"]].get(i) for i in inst_info.get(attr["attr_id"], [])])
                    )
                    continue

                _value = inst_info.get(attr["attr_id"])
                if attr["attr_type"] == ENUM:
                    _value = enum_field_dict[attr["attr_id"]].get(_value)
                sheet_data.append(_value)
            workbook.active.append(sheet_data)
        return self.return_bytesio(workbook)
